import pandas
import collections
from openpyxl import load_workbook

GENDER_MALE = 'м'
GENDER_FEMALE = 'ж'

def list_by_month(*args):
    #key, sum, data
    """Согласование окончаний существительных
           :param key: ключ списка
           :param sum: общая сумма (0-й элемент)
           :param dat: дата посещения
           :param list: список
    """
    # v1 = el['Браузер']  # ключ - нащвание браузерв
    # v3 = brawser_list[v1]  # сумма - количество посещений

    key = args[0]
    sum = args[1]
    dat = args[2]
    list = args[3]

    # получение месяца
    d = int(dat.date().month)

    if key in list.keys():
        mes = list[key]
    else:
        mes = {0: sum, 1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0, 12: 0}

    mes[d] += 1
    list[key] = mes

    return list

# Активируем выходной файл-шаблон для отчета
dest_filename = 'report.xlsx'
wb = load_workbook(filename=dest_filename, data_only = True)

sheet = wb['Лист1']
################################

# получить имя первого листа
name = pandas.ExcelFile('logs.xlsx', engine="openpyxl").sheet_names[0]
# прочитать 1-й лист книги
rb = pandas.read_excel('logs.xlsx', engine="openpyxl", sheet_name=name)


# получить словарь всего листа Excel
list_prev = rb.to_dict(orient='records')

#получить коллекцию популярных браузеров
list_brawser = rb['Браузер'].tolist()
brawser_counter = collections.Counter(list_brawser)
brawsers = brawser_counter.most_common(7)
brawser_list = {}

for k, v in brawsers:
    brawser_list[k] = v

brawser_dict = {}
 # собрать словарь браузеров
for el in list_prev:
    #выбрать только записи, которые относятся к 7 популярным браузерам
    if el['Браузер'] in brawser_list.keys():
        
        vn = el['Браузер'] # ключ - нащвание браузерв
        vs = brawser_list[vn] # сумма - количество посещений
        
        # получение месяца
        vd = el['Дата посещения']

        brawser_dict = list_by_month(vn, vs, vd, brawser_dict)

# отсортировать по популярности:
sort_brawsers = {}
for k, v in brawsers:
    sort_brawsers[k] = brawser_dict[k]

print('Список популярных браузеров: ')
print(sort_brawsers)

# вывести в отчет
sum_all = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
j = 0
for k, v in sort_brawsers.items():
    key = k
    sheet.cell(column=1, row=5 + j, value= key)
    list_val = v
    i = 0
    for km, kv in list_val.items():
        sheet.cell(column=2 + i, row=5 + j, value=kv)
        sum_all[km] = sum_all[km] +  kv
        i += 1

    j += 1

i = 0
for sum in sum_all:
    sheet.cell(column=2 + i, row=5 + j, value=sum)
    i += 1

#####################################################################################################
#получить коллекцию популярных товаров:
list_product_prev = rb['Купленные товары'].tolist()

#Получаем список всех продуктов
list_product = []

for el in list_product_prev:
    r_list = el.split(',')
    for el_in in r_list:
        if str(el_in).find('вариант') < 0:
            list_product.append(el_in.strip())


#получить коллекцию популярных товаров
product_counter = collections.Counter(list_product)
products =product_counter.most_common(7)

product_list = {}
for k, v in products:
    product_list[k] = v

# ТОВАРЫ:
product_dict = {}

list_male = [] #список тоаров мужских
list_female = [] #список тоаров женских

# собрать словарь товаров
for el in list_prev:

    gender = el['Пол']
    records = el['Купленные товары'].split(',')

    for el_in in records:
        if str(el_in).find('вариант') < 0:
            if gender == GENDER_MALE:
                list_male.append(el_in)
            elif gender == GENDER_FEMALE:
                list_female.append(el_in)



    i = 0
    #выбрать только записи, которые относятся к 7 популярным товарам
    while i < len(records):

        if records[i] in product_list.keys():#if records[i] in product_list.keys():

            vn = records[i] # ключ - название товара
            vs = product_list[records[i]]# сумма количество продаж
            # получение даты
            vd = el['Дата посещения']

            product_dict = list_by_month(vn, vs, vd, product_dict)
        i += 1


# отсортировать по популярности:
sort_products = {}
for k, v in products:
    sort_products[k] = product_dict[k]


print('Список популярных товаров:')
print(sort_products)
# вывести в отчет
sum_all = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
j = 0
for k, v in sort_products.items():
    key = k
    sheet.cell(column=1, row=19 + j, value= key)
    list_val = v
    i = 0
    for km, kv in list_val.items():
        sheet.cell(column=2 + i, row=19 + j, value=kv)
        sum_all[km] = sum_all[km] +  kv
        i += 1

    j += 1

i = 0
for sum in sum_all:
    sheet.cell(column=2 + i, row=19 + j, value=sum)
    i += 1


# заполнить раздел Предпочтения
male_counter = collections.Counter(list_male)
female_counter = collections.Counter(list_female)

max_male_counter_val =  male_counter.most_common(1)[0][0]
max_female_counter_val =  female_counter.most_common(1)[0][0]

min_male_counter_val = male_counter.most_common()[-1][0]
min_female_counter_val =  female_counter.most_common()[-1][0]

print('max male:', max_male_counter_val) # 'Мешок для обуви G9807 (черный)', 101
print('max female:', max_female_counter_val) # 'Мешок для обуви G9807 (черный)', 91

print('min male:', min_male_counter_val) #  'Браслет Xiaomi Mi Band 5 (CN) черный', 22
print('min female:', min_female_counter_val) # 'ArtSpace Набор обложек для дневников и тетрадей  208х346 мм', 21

# вывести в отчет
sheet['B31'] = max_male_counter_val
sheet['B32'] = max_female_counter_val

sheet['B33'] = min_male_counter_val
sheet['B34'] = min_female_counter_val

# сохранить отчет
wb.save(filename = dest_filename)










